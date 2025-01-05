from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Load the Excel file
file_path = "Transmission Loss (db).xlsx"  # Replace with your actual file path
workbook = load_workbook(file_path, data_only=True)  # `data_only=True` extracts computed values for cells with formulas

# Specify the sheets to read and their corresponding line colors
sheets_to_plot = {
    "SPI-A-Av": "b",
    "SPI-B-Av": "g",
    "SCA-B1-Av": "darkorange",
    "SCA-B2-Av": "navy",
    "E3DCP-Concrete-Av": "violet",
    "E3DCP-Cast Pattern-Av": "crimson",
    "E3DCP-Clay-Av": "gold",
    "Shotcrete-A-Av": "aqua",
    "Shotcrete-B-Av": "lime"
}

# Frequencies corresponding to the transmission loss values
frequencies = [62.5, 125, 250, 500, 1000, 1800]  # Frequencies in Hz

# Prepare the plot
plt.figure(figsize=(12, 8))

# Iterate through the specified sheets
for sheet_name, color in sheets_to_plot.items():
    # Load the current sheet
    sheet = workbook[sheet_name]
    
    # Extract transmission loss data from the sheet
    transmission_loss = []
    for row in sheet.iter_rows(min_row=2, max_row=7, min_col=2, max_col=2, values_only=True):
        if row[0] is not None:  # Ensure the value is not None
            transmission_loss.append(float(row[0]))  # Convert to float for plotting
    
    # Plot the data with the specified color
    # Plot the line with the specified color and markers at each frequency point
    plt.plot(frequencies, transmission_loss, label=sheet_name, color=color, marker='o', markersize=6)

# Customize the plot
plt.title("Average Sound Transmission Loss of AM Materials", fontsize=16)
plt.xlabel("Frequency (Hz)", fontsize=14)
plt.ylabel("Transmission Loss (dB)", fontsize=14)
plt.xscale('log')  # Logarithmic scale for frequency

# Remove the grid lines completely
plt.grid(False)

# Set the x-axis ticks to explicitly show the specified frequencies
plt.xticks(frequencies, labels=[f"{freq} Hz" for freq in frequencies], fontsize=10)

# Manually set the y-axis range and ticks
plt.ylim(5, 55)  # Set y-axis limits from 0 to 60 dB
plt.yticks(range(5, 56, 5), fontsize=10)  # Set ticks at intervals of 5 dB

# Move the legend below the graph
plt.legend(title="Materials", loc='upper center', bbox_to_anchor=(0.5, -0.10), ncol=4, fontsize=14)

# Adjust layout
plt.tight_layout()

# Show the plot
output_file = "DTL Comparison.png"
plt.savefig(output_file, dpi=300)
plt.show()

print(f"Plot saved as {output_file}")