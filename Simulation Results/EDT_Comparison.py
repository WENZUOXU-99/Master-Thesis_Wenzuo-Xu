from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

# Load the Excel file
file_path = "EDT.xlsx"  # Replace with your actual file path
workbook = load_workbook(file_path, data_only=True)

# Specify the sheets to read and their corresponding line colors
sheets_to_plot = {
    "SPI-A": "b",
    "SPI-B": "g",
    "SCA-B1": "lime",
    "SCA-B2": "crimson",
    "E3DCP-Concrete": "k",
    "E3DCP-Cast Pattern": "darkgray",
    "E3DCP-Clay": "darkorange",
    "Shotcrete-A": "aqua",
    "Shotcrete-B": "violet"
}

# Frequencies corresponding to the transmission loss values
frequencies = np.array([62.5, 125, 250, 500, 1000, 1800])  # Frequencies in Hz

# Initialize a dictionary to store transmission loss values for each frequency
frequency_data = {freq: [] for freq in frequencies}

# Prepare the plot with increased height
plt.figure(figsize=(12, 16))  # Increase the figure height to create more vertical space

# Thread width adjustment factor
thread_factor = 1.15  # Small factor to adjust width in log scale

# Iterate through the specified sheets
for sheet_name, color in sheets_to_plot.items():
    # Load the current sheet
    sheet = workbook[sheet_name]

    # Extract transmission loss data from the sheet
    transmission_loss = []
    for row in sheet.iter_rows(min_row=2, max_row=7, min_col=2, max_col=2, values_only=True):
        if row[0] is not None:  # Ensure the value is not None
            transmission_loss.append(float(row[0]))  # Convert to float for plotting

    # Add data to the frequency_data dictionary
    for freq, tl in zip(frequencies, transmission_loss):
        frequency_data[freq].append(tl)

    # Plot the data as threads (vertical lines)
    for freq, tl in zip(frequencies, transmission_loss):
        plt.plot(
            [freq / thread_factor, freq * thread_factor],  # Adjusted thread width
            [tl, tl],
            color=color,
            lw=2,
            label=sheet_name if freq == frequencies[0] else ""  # Add legend label only once
        )

# Annotate the highest and lowest values
for freq, values in frequency_data.items():
    max_value = max(values)
    min_value = min(values)
    
    # Add the max value above the thread
    plt.text(
        freq, max_value + 0.005,  # Slight offset above the thread
        f"{max_value:.3f}",
        ha='center', va='bottom', fontsize=10, color="black"
    )
    
    # Add the min value below the thread
    plt.text(
        freq, min_value - 0.005,  # Slight offset below the thread
        f"{min_value:.3f}",
        ha='center', va='top', fontsize=10, color="black"
    )

# Customize the plot
plt.title("Early Decay Time (EDT) with different AM Materials as Ceiling", fontsize=18)
plt.xlabel("Frequency (Hz)", fontsize=16)
plt.ylabel("EDT in s", fontsize=16)
plt.xscale('log')  # Logarithmic scale for frequency

# Set the x-axis ticks to explicitly show the specified frequencies
plt.xticks(frequencies, labels=[f"{freq} Hz" for freq in frequencies], fontsize=12)

# Manually set the y-axis range and ticks
plt.ylim(0.275, 0.55)  # Set y-axis limits
plt.yticks(np.arange(0.275, 0.60, 0.025), fontsize=12)  # Set ticks at intervals of 0.025

# Move the legend below the graph
plt.legend(title="Materials", loc='upper center', bbox_to_anchor=(0.5, -0.10), ncol=3, fontsize=16)

# Adjust layout to ensure adequate spacing
plt.tight_layout(pad=3.0)  # Add padding around the plot

# Show the plot
output_file = "EDT_Threads_Narrow_YAxis_Annotated.png"
plt.savefig(output_file, dpi=300)
plt.show()

print(f"Plot saved as {output_file}")

